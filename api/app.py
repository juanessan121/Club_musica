import base64
import hashlib
import hmac
import html as _html
import logging
import os
import re as _re
import uuid
from datetime import date, datetime, timedelta
from functools import wraps
import string
import secrets
import smtplib

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("club_musica")
import io
import threading
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pymysql
from dbutils.pooled_db import PooledDB
from helpers import normalize_level, normalize_role, normalize_socio_estado, is_valid_operating_date, require_fields
import requests
import jwt
from flask import Flask, jsonify, request, Response
from flask_cors import CORS
from flask_talisman import Talisman

try:
    import bcrypt
except ImportError:
    bcrypt = None


app = Flask(__name__)

# ── Cookies de sesión seguras ────────────────────────────────────────────────
app.config.update(
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Strict",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=int(os.environ.get("JWT_EXPIRY_HOURS", "24"))),
)

# ── Flask-Talisman: HSTS, CSP, X-Frame-Options, Referrer-Policy ─────────────
# force_https=False porque nginx ya hace la terminación TLS y el redirect HTTP→HTTPS.
# Los headers HSTS llegan al browser igualmente porque nginx no los filtra.
_CSP = {
    "default-src": ["'self'"],
    "script-src":  ["'self'", "'unsafe-inline'"],
    "style-src":   ["'self'", "'unsafe-inline'"],
    "img-src":     ["'self'", "data:", "https:"],
    "font-src":    ["'self'", "data:"],
    "connect-src": ["'self'", "wss:"],
    "frame-ancestors": ["'none'"],
}
Talisman(
    app,
    force_https=False,
    strict_transport_security=True,
    strict_transport_security_max_age=63072000,
    strict_transport_security_include_subdomains=True,
    content_security_policy=_CSP,
    referrer_policy="strict-origin-when-cross-origin",
    frame_options="DENY",
    x_content_type_options=True,
)

# ── Configuración de seguridad ───────────────────────────────────────────────
ALLOWED_ORIGINS = os.environ.get(
    "CORS_ALLOWED_ORIGINS",
    "http://localhost:3001,http://localhost:8088,http://127.0.0.1:3001"
).split(",")
CORS(app, origins=ALLOWED_ORIGINS, supports_credentials=True)

def read_secret(env_var: str, secret_name: str | None = None, default: str = "") -> str:
    """Lee un secreto desde /run/secrets/ (Docker Secrets) y cae en la variable de entorno.
    Esto permite usar Docker Secrets en producción sin cambiar el código."""
    if secret_name:
        try:
            with open(f"/run/secrets/{secret_name}") as _f:
                return _f.read().strip()
        except FileNotFoundError:
            pass
    return os.environ.get(env_var, default)


DB_HOST     = os.environ.get("DB_HOST", "db-mariadb")
DB_USER     = os.environ.get("DB_USER", "clubmusica")
DB_PASSWORD = read_secret("DB_PASSWORD", "db_password")
DB_NAME     = os.environ.get("DB_NAME", "club_musica")
WHATSAPP_BRIDGE_URL = os.environ.get("WHATSAPP_BRIDGE_URL", "http://whatsapp-bridge:3002")
DEFAULT_PASSWORD    = os.environ.get("DEFAULT_USER_PASSWORD", "Musica2026!")

JWT_SECRET    = read_secret("JWT_SECRET", "jwt_secret", "super-secret-key-2026")
JWT_ALGORITHM = "HS256"
JWT_EXPIRY_HOURS = int(os.environ.get("JWT_EXPIRY_HOURS", "24"))

SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER', 'proyectoserror404@gmail.com')
SMTP_PASS = os.environ.get('SMTP_PASS', '')
SMTP_FROM = os.environ.get('SMTP_FROM', 'proyectoserror404@gmail.com')

SCHEMA_READY = False

# Advertencia de seguridad si JWT_SECRET es débil
if len(JWT_SECRET) < 32 or JWT_SECRET in ("super-secret-key-2026", "cambia-este-secreto-en-produccion"):
    logger.warning("[SECURITY WARNING] JWT_SECRET is weak or default. Set a strong random secret (min 32 chars).")

# ── Blacklist de tokens revocados (logout) — persistida en BD ────────────────
# Ya no usamos un set en RAM: los tokens revocados sobreviven reinicios y son
# compartidos entre todos los workers de gunicorn.

def _is_token_revoked(jti: str) -> bool:
    """Devuelve True si el jti está en TOKEN_BLACKLIST y aún no ha expirado."""
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT 1 FROM TOKEN_BLACKLIST WHERE jti = %s AND expires_at > %s",
                (jti, datetime.utcnow()),
            )
            return cursor.fetchone() is not None
    except Exception:
        logger.error("TOKEN_BLACKLIST lookup error", exc_info=True)
        return False  # fail-open: no bloquear usuarios ante un fallo de BD
    finally:
        conn.close()


def _revoke_token(jti: str, expires_at: datetime) -> None:
    """Inserta un jti en TOKEN_BLACKLIST. INSERT IGNORE tolera duplicados."""
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute(
                "INSERT IGNORE INTO TOKEN_BLACKLIST (jti, expires_at) VALUES (%s, %s)",
                (jti, expires_at),
            )
        conn.commit()
    except Exception:
        logger.error("TOKEN_BLACKLIST insert error", exc_info=True)
    finally:
        conn.close()


def _cleanup_token_blacklist() -> None:
    """Elimina entradas expiradas de TOKEN_BLACKLIST. Llamado por el scheduler."""
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute(
                "DELETE FROM TOKEN_BLACKLIST WHERE expires_at < %s",
                (datetime.utcnow(),),
            )
            deleted = cursor.rowcount
        conn.commit()
        if deleted:
            logger.info("TOKEN_BLACKLIST cleanup: %d entradas expiradas eliminadas", deleted)
    except Exception:
        logger.error("TOKEN_BLACKLIST cleanup error", exc_info=True)
    finally:
        conn.close()

# ── Security headers adicionales ────────────────────────────────────────────
# Talisman ya inyecta HSTS, CSP, X-Frame-Options, X-Content-Type-Options y
# Referrer-Policy. Solo necesitamos eliminar el header Server para no revelar
# información del stack.
@app.after_request
def add_security_headers(response):
    response.headers.pop("Server", None)
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    return response

def get_user_from_token():
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return None
    token = auth_header.split(" ", 1)[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.InvalidTokenError:
        return None
    jti = payload.get("jti")
    if jti and _is_token_revoked(jti):
        return None
    return payload

def _verify_user_in_db(payload):
    """Verifica que el usuario del token existe, está activo y tiene el rol correcto en BD."""
    user_id = payload.get("id")
    if not user_id:
        return None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT id, rol, estado FROM SOCIO WHERE id = %s",
                (user_id,),
            )
            row = cursor.fetchone()
        conn.close()
    except Exception as e:
        logger.error("[AUTH] _verify_user_in_db error: %s", e)
        return None
    if not row or row["estado"] != "ACTIVO":
        return None
    # Sobrescribe el rol con el valor real de la BD (no el del token)
    payload["rol"] = row["rol"]
    return payload

def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        payload = get_user_from_token()
        if not payload:
            return error("Token inválido o expirado", 401)
        user = _verify_user_in_db(payload)
        if not user:
            return error("Sesión inválida o cuenta inactiva", 401)
        request.user = user
        return f(*args, **kwargs)
    return decorated

def require_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        payload = get_user_from_token()
        if not payload:
            return error("Token inválido o expirado", 401)
        user = _verify_user_in_db(payload)
        if not user:
            return error("Sesión inválida o cuenta inactiva", 401)
        if user.get("rol") != "ADMIN":
            return error("No tienes permisos para realizar esta acción", 403)
        request.user = user
        return f(*args, **kwargs)
    return decorated


_db_pool = PooledDB(
    creator=pymysql,
    mincached=0,
    maxcached=5,
    maxconnections=10,
    blocking=True,
    ping=2,
    host=DB_HOST,
    user=DB_USER,
    password=DB_PASSWORD,
    database=DB_NAME,
    cursorclass=pymysql.cursors.DictCursor,
    autocommit=False,
)

def get_db_connection():
    return _db_pool.connection()


def ok(data=None, status=200):
    return jsonify({} if data is None else data), status


def error(message, status=400):
    return jsonify({"error": message}), status


def parse_json():
    return request.get_json(silent=True) or {}


def _build_safe_set(updates, allowed_cols):
    """Construye la cláusula SET de un UPDATE validando cada columna contra un allowlist.
    Lanza ValueError si algún nombre de columna no está en allowed_cols."""
    for col, _ in updates:
        if col not in allowed_cols:
            raise ValueError(f"Columna no permitida: {col}")
    return ", ".join(f"{col} = %s" for col, _ in updates)  # nosec B608


def serialize_row(row):
    if not row:
        return row
    serialized = {}
    for key, value in row.items():
        if isinstance(value, datetime):
            serialized[key] = value.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(value, date):
            serialized[key] = value.isoformat()
        else:
            serialized[key] = value
    return serialized


def serialize_rows(rows):
    return [serialize_row(row) for row in rows]


def parse_datetime(value, field_name):
    if isinstance(value, datetime):
        return value
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            continue
    raise ValueError(f"{field_name} debe tener formato datetime válido")


def parse_date(value, field_name):
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    raw = str(value)
    if "T" in raw:
        raw = raw.split("T", 1)[0]
    if " " in raw:
        raw = raw.split(" ", 1)[0]
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(f"{field_name} debe tener formato YYYY-MM-DD") from exc


def hash_password(password):
    if bcrypt:
        return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt(rounds=13)).decode("utf-8")

    salt = os.urandom(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 600000)
    return "pbkdf2_sha256$%s$%s" % (
        base64.b64encode(salt).decode("ascii"),
        base64.b64encode(digest).decode("ascii"),
    )


def verify_password(password, stored_hash):
    if not stored_hash:
        return False
    if stored_hash.startswith("$2") and bcrypt:
        return bcrypt.checkpw(password.encode("utf-8"), stored_hash.encode("utf-8"))
    if stored_hash.startswith("pbkdf2_sha256$"):
        _, salt_b64, digest_b64 = stored_hash.split("$", 2)
        salt = base64.b64decode(salt_b64)
        expected = base64.b64decode(digest_b64)
        actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 600000)
        return hmac.compare_digest(actual, expected)
    return False


def table_exists(cursor, table_name):
    cursor.execute("SHOW TABLES LIKE %s", (table_name,))
    return cursor.fetchone() is not None


_ALLOWED_TABLES = {"SOCIO", "RESERVA", "PRESTAMO", "INSTRUMENTO", "SALA", "MULTA", "EVENTO"}

def column_names(cursor, table_name):
    if table_name not in _ALLOWED_TABLES:
        raise ValueError(f"Tabla no permitida: {table_name}")
    cursor.execute(f"SHOW COLUMNS FROM {table_name}")
    return {row["Field"] for row in cursor.fetchall()}


def trigger_exists(cursor, trigger_name):
    cursor.execute(
        """
        SELECT TRIGGER_NAME
        FROM information_schema.TRIGGERS
        WHERE TRIGGER_SCHEMA = DATABASE() AND TRIGGER_NAME = %s
        """,
        (trigger_name,),
    )
    return cursor.fetchone() is not None


def ensure_database_schema():
    """Small compatibility pass for databases created before auth columns existed."""
    global SCHEMA_READY
    if SCHEMA_READY:
        return

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            if not table_exists(cursor, "SOCIO"):
                raise RuntimeError("La base de datos no está inicializada con schema_musica.sql")

            socio_cols = column_names(cursor, "SOCIO")
            if "password_hash" not in socio_cols:
                cursor.execute("ALTER TABLE SOCIO ADD COLUMN password_hash VARCHAR(255) NOT NULL DEFAULT '' AFTER telefono")
            if "password_salt" not in socio_cols:
                cursor.execute("ALTER TABLE SOCIO ADD COLUMN password_salt VARCHAR(255) NOT NULL DEFAULT '' AFTER password_hash")
            if "rol" not in socio_cols:
                cursor.execute("ALTER TABLE SOCIO ADD COLUMN rol VARCHAR(20) NOT NULL DEFAULT 'SOCIO' AFTER nivel_habilidad")
            cursor.execute("UPDATE SOCIO SET rol = 'ADMIN' WHERE email = 'juan.sandoval@pucesa.edu.ec'")
            cursor.execute("UPDATE SOCIO SET rol = 'SOCIO' WHERE rol IS NULL OR rol = ''")

            # Add notificado_1h column to RESERVA if missing
            reserva_cols = column_names(cursor, 'RESERVA')
            if 'notificado_1h' not in reserva_cols:
                cursor.execute('ALTER TABLE RESERVA ADD COLUMN notificado_1h TINYINT(1) DEFAULT 0')
            # Add avatar_url to SOCIO if missing
            socio_cols2 = column_names(cursor, 'SOCIO')
            if 'avatar_url' not in socio_cols2:
                cursor.execute('ALTER TABLE SOCIO ADD COLUMN avatar_url TEXT NULL')

            # MULTA table
            if not table_exists(cursor, 'MULTA'):
                cursor.execute("""
                    CREATE TABLE MULTA (
                        id INT PRIMARY KEY AUTO_INCREMENT,
                        socio_id INT NOT NULL,
                        prestamo_id INT NULL,
                        reserva_id INT NULL,
                        monto DECIMAL(8,2) NOT NULL DEFAULT 0.00,
                        motivo VARCHAR(255) NOT NULL,
                        estado VARCHAR(20) NOT NULL DEFAULT 'PENDIENTE',
                        fecha_creacion DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        fecha_pago DATETIME NULL,
                        FOREIGN KEY (socio_id) REFERENCES SOCIO(id)
                    ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
                """)

            # PRESTAMO table changes
            prestamo_cols = column_names(cursor, 'PRESTAMO')
            if 'motivo' not in prestamo_cols:
                cursor.execute('ALTER TABLE PRESTAMO ADD COLUMN motivo VARCHAR(255) NULL AFTER instrumento_id')
            
            try:
                cursor.execute('ALTER TABLE PRESTAMO MODIFY COLUMN fecha_salida DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP')
                cursor.execute('ALTER TABLE PRESTAMO MODIFY COLUMN fecha_limite DATETIME NOT NULL')
                cursor.execute('ALTER TABLE PRESTAMO MODIFY COLUMN fecha_devolucion DATETIME NULL')
            except Exception as e:
                print("Nota: No se pudo alterar PRESTAMO a DATETIME:", e)

            # Fix: instrumentos marcados PRESTADO sin préstamo activo → DISPONIBLE
            cursor.execute("""
                UPDATE INSTRUMENTO i
                SET i.estado = 'DISPONIBLE', i.modificado_por = 'MIGRATION'
                WHERE i.estado = 'PRESTADO'
                AND NOT EXISTS (
                    SELECT 1 FROM PRESTAMO p
                    WHERE p.instrumento_id = i.id
                    AND p.estado = 'ACTIVO'
                    AND p.eliminado_en IS NULL
                )
            """)
            if cursor.rowcount > 0:
                print(f"[MIGRATION] Corregidos {cursor.rowcount} instrumentos PRESTADO sin préstamo activo.")

            conn.commit()
            SCHEMA_READY = True
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


@app.before_request
def migrate_existing_database():
    if request.path == "/api/health":
        return None
    ensure_database_schema()
    return None


def _now_local():
    """Hora actual en Ecuador (UTC-5) como datetime naive, independiente del timezone del servidor."""
    return datetime.utcnow() - timedelta(hours=5)

def within_operating_hours():
    """Lunes a sábado hasta las 12:00 hora Ecuador. Sábado tarde y domingo bloqueados."""
    now = _now_local()
    weekday = now.weekday()  # 0=Lun … 5=Sáb, 6=Dom
    if weekday == 6:
        return False
    if weekday == 5 and now.hour >= 12:
        return False
    return True

OPERATING_HOURS_MSG = (
    "Las reservas y préstamos solo se pueden gestionar "
    "de lunes a sábado hasta las 12:00 del mediodía."
)
OPERATING_DATE_MSG = (
    "La fecha seleccionada no es válida: no se permiten reservas ni préstamos "
    "en domingo ni sábado a partir de las 12:00."
)


def normalize_sala_tipo(value):
    allowed = {"CUBICULO", "SALON_ACUSTICO", "ESTUDIO"}
    normalized = str(value or "CUBICULO").upper()
    return normalized if normalized in allowed else "CUBICULO"


def normalize_sala_estado(value):
    aliases = {"DISPONIBLE": "ACTIVA", "CLAUSURADA": "INACTIVA"}
    normalized = aliases.get(str(value or "ACTIVA").upper(), str(value or "ACTIVA").upper())
    return normalized if normalized in {"ACTIVA", "MANTENIMIENTO", "INACTIVA"} else "ACTIVA"


def normalize_instrumento_estado(value):
    aliases = {
        "EXCELENTE": "DISPONIBLE",
        "BUENO": "DISPONIBLE",
        "REGULAR": "DISPONIBLE",
        "DAÑADO": "MANTENIMIENTO",
        "DANADO": "MANTENIMIENTO",
        "EN_MANTENIMIENTO": "MANTENIMIENTO",
    }
    normalized = aliases.get(str(value or "DISPONIBLE").upper(), str(value or "DISPONIBLE").upper())
    return normalized if normalized in {"DISPONIBLE", "PRESTADO", "MANTENIMIENTO", "BAJA"} else "DISPONIBLE"


def normalize_reserva_estado(value):
    aliases = {"CONFIRMADA": "CONFIRMADA", "PENDIENTE": "CONFIRMADA"}
    normalized = aliases.get(str(value or "CONFIRMADA").upper(), str(value or "CONFIRMADA").upper())
    allowed = {"CONFIRMADA", "CANCELADA", "COMPLETADA", "INASISTENCIA", "REPROGRAMADA"}
    return normalized if normalized in allowed else "CONFIRMADA"


def normalize_prestamo_estado(value):
    normalized = str(value or "ACTIVO").upper()
    return normalized if normalized in {"ACTIVO", "DEVUELTO", "VENCIDO"} else "ACTIVO"


def socio_public(row):
    row = serialize_row(row)
    return {
        "id": row["id"],
        "nombre_completo": row["nombre"],
        "email_institucional": row["email"],
        "telefono_whatsapp": row.get("telefono"),
        "nivel_habilidad": row.get("nivel_habilidad"),
        "rol": row.get("rol", "SOCIO"),
        "estado": row.get("estado"),
        "fecha_registro": row.get("fecha_registro"),
    }


def get_or_create_tipo_instrumento(cursor, nombre):
    nombre = (nombre or "Otro").strip()
    cursor.execute("SELECT id FROM TIPO_INSTRUMENTO WHERE LOWER(nombre) = LOWER(%s)", (nombre,))
    found = cursor.fetchone()
    if found:
        return found["id"]
    cursor.execute("INSERT INTO TIPO_INSTRUMENTO (nombre) VALUES (%s)", (nombre,))
    return cursor.lastrowid


def notify_whatsapp(number, message):
    if not number:
        return False
    try:
        response = requests.post(
            f"{WHATSAPP_BRIDGE_URL}/send",
            json={"number": number, "message": message},
            timeout=3,
        )
        return response.status_code < 500
    except requests.RequestException:
        return False


def send_email_async(to_email, subject, html_body):
    """Send email in background thread to avoid blocking the API."""
    def _send():
        if not SMTP_USER or not SMTP_PASS:
            print(f"[EMAIL SIMULADO] Para: {to_email} | Asunto: {subject}")
            return
        try:
            msg = MIMEMultipart('alternative')
            msg['Subject'] = subject
            msg['From'] = f'Club de M\u00fasica PUCESA <{SMTP_FROM}>'
            msg['To'] = to_email
            msg.attach(MIMEText(html_body, 'html', 'utf-8'))
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
                server.ehlo()
                server.starttls()
                server.login(SMTP_USER, SMTP_PASS)
                server.sendmail(SMTP_FROM, [to_email], msg.as_string())
            print(f"[EMAIL OK] Para: {to_email} | Asunto: {subject}")
        except Exception as e:
            print(f"[EMAIL ERROR] {e}")
    threading.Thread(target=_send, daemon=True).start()


def build_email_reserva_confirmada(nombre, sala_nombre, fecha_inicio, fecha_fin, reserva_id):
    nombre, sala_nombre = _html.escape(str(nombre)), _html.escape(str(sala_nombre))
    fecha_inicio, fecha_fin = _html.escape(str(fecha_inicio)), _html.escape(str(fecha_fin))
    return f"""
<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Reserva Confirmada</title>
</head>
<body style="margin:0;padding:0;background:#f0f4f8;font-family:'Segoe UI',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:40px 0;">
    <tr><td align="center">
      <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.10);">
        <!-- Header -->
        <tr>
          <td style="background:linear-gradient(135deg,#116a7b 0%,#0b4d59 100%);padding:40px 48px 32px;">
            <table width="100%">
              <tr>
                <td>
                  <p style="margin:0;color:rgba(255,255,255,0.7);font-size:13px;letter-spacing:2px;text-transform:uppercase;">Club de M&uacute;sica</p>
                  <h1 style="margin:8px 0 0;color:#ffffff;font-size:28px;font-weight:800;letter-spacing:-0.5px;">&#9989; Reserva Confirmada</h1>
                </td>
                <td align="right">
                  <div style="background:rgba(255,255,255,0.15);border-radius:12px;padding:12px 18px;display:inline-block;">
                    <p style="margin:0;color:white;font-size:11px;opacity:0.8;">ID DE RESERVA</p>
                    <p style="margin:4px 0 0;color:white;font-size:22px;font-weight:800;">#{reserva_id}</p>
                  </div>
                </td>
              </tr>
            </table>
          </td>
        </tr>
        <!-- Greeting -->
        <tr>
          <td style="padding:36px 48px 0;">
            <p style="margin:0;font-size:18px;color:#172033;">Hola, <strong>{nombre}</strong> &#128075;</p>
            <p style="margin:12px 0 0;font-size:15px;color:#647084;line-height:1.6;">Tu reserva de sala ha sido <strong style="color:#116a7b;">confirmada exitosamente</strong>. Aqu&iacute; est&aacute;n los detalles:</p>
          </td>
        </tr>
        <!-- Details card -->
        <tr>
          <td style="padding:24px 48px;">
            <div style="background:#f8fafc;border-radius:12px;padding:24px;border:1.5px solid #dbe1ea;">
              <table width="100%" cellpadding="0" cellspacing="0">
                <tr>
                  <td style="padding:10px 0;border-bottom:1px solid #edf2f7;">
                    <span style="font-size:13px;color:#647084;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">&#127968; Sala</span>
                    <p style="margin:4px 0 0;font-size:17px;font-weight:700;color:#172033;">{sala_nombre}</p>
                  </td>
                </tr>
                <tr>
                  <td style="padding:10px 0;border-bottom:1px solid #edf2f7;">
                    <span style="font-size:13px;color:#647084;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">&#128197; Inicio</span>
                    <p style="margin:4px 0 0;font-size:17px;font-weight:700;color:#172033;">{fecha_inicio}</p>
                  </td>
                </tr>
                <tr>
                  <td style="padding:10px 0;">
                    <span style="font-size:13px;color:#647084;font-weight:600;text-transform:uppercase;letter-spacing:0.5px;">&#128276; Fin</span>
                    <p style="margin:4px 0 0;font-size:17px;font-weight:700;color:#172033;">{fecha_fin}</p>
                  </td>
                </tr>
              </table>
            </div>
          </td>
        </tr>
        <!-- Rules reminder -->
        <tr>
          <td style="padding:0 48px 24px;">
            <div style="background:#fff7ed;border-radius:12px;padding:20px 24px;border-left:4px solid #f59e0b;">
              <p style="margin:0 0 8px;font-size:13px;font-weight:700;color:#92400e;">&#128203; RECUERDA LAS REGLAS</p>
              <ul style="margin:0;padding:0 0 0 16px;font-size:13px;color:#78350f;line-height:1.8;">
                <li>Llega puntual &mdash; la sala se libera autom&aacute;ticamente al terminar tu turno.</li>
                <li>Deja la sala en las mismas condiciones que la encontraste.</li>
                <li>En caso de no asistir, cancela con al menos 30 minutos de anticipaci&oacute;n.</li>
              </ul>
            </div>
          </td>
        </tr>
        <!-- Footer -->
        <tr>
          <td style="background:#f8fafc;padding:24px 48px;border-top:1.5px solid #dbe1ea;">
            <p style="margin:0;font-size:13px;color:#647084;text-align:center;">Club de M&uacute;sica &middot; PUCESA &middot; Este correo fue generado autom&aacute;ticamente.</p>
          </td>
        </tr>
      </table>
    </td></tr>
  </table>
</body>
</html>
    """


def build_email_reserva_cancelada(nombre, sala_nombre, fecha_inicio, reserva_id):
    nombre, sala_nombre, fecha_inicio = _html.escape(str(nombre)), _html.escape(str(sala_nombre)), _html.escape(str(fecha_inicio))
    return f"""
<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><title>Reserva Cancelada</title></head>
<body style="margin:0;padding:0;background:#f0f4f8;font-family:'Segoe UI',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:40px 0;">
    <tr><td align="center">
      <table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.10);">
        <tr>
          <td style="background:linear-gradient(135deg,#b42318 0%,#7f1d1d 100%);padding:40px 48px 32px;">
            <p style="margin:0;color:rgba(255,255,255,0.7);font-size:13px;letter-spacing:2px;text-transform:uppercase;">Club de M&uacute;sica</p>
            <h1 style="margin:8px 0 0;color:#ffffff;font-size:28px;font-weight:800;">&#10060; Reserva Cancelada</h1>
          </td>
        </tr>
        <tr><td style="padding:36px 48px 24px;">
          <p style="margin:0;font-size:16px;color:#172033;">Hola <strong>{nombre}</strong>, tu reserva <strong>#{reserva_id}</strong> de la sala <strong style="color:#b42318;">{sala_nombre}</strong> programada para el <strong>{fecha_inicio}</strong> ha sido cancelada.</p>
          <p style="margin:16px 0 0;font-size:14px;color:#647084;">Si crees que esto es un error o deseas hacer una nueva reserva, ingresa al sistema.</p>
          <div style="margin:24px 0 0;background:#fff5f5;border-radius:10px;padding:16px 20px;border-left:4px solid #b42318;">
            <p style="margin:0;font-size:13px;color:#7f1d1d;">&#128161; Recuerda que puedes hacer una nueva reserva en cualquier momento desde el sistema.</p>
          </div>
        </td></tr>
        <tr><td style="background:#f8fafc;padding:20px 48px;border-top:1.5px solid #dbe1ea;">
          <p style="margin:0;font-size:13px;color:#647084;text-align:center;">Club de M&uacute;sica &middot; PUCESA</p>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body>
</html>
    """


def reservation_overlap(cursor, sala_id, fecha_inicio, fecha_fin, exclude_id=None):
    params = [sala_id, fecha_fin, fecha_inicio]
    sql = """
        SELECT id
        FROM RESERVA
        WHERE sala_id = %s
          AND eliminado_en IS NULL
          AND estado IN ('CONFIRMADA', 'REPROGRAMADA')
          AND fecha_inicio < %s
          AND fecha_fin > %s
    """
    if exclude_id:
        sql += " AND id <> %s"
        params.append(exclude_id)
    cursor.execute(sql, params)
    return cursor.fetchone() is not None


@app.route("/api/health", methods=["GET"])
def health():
    return ok({"status": "healthy", "service": "club-musica-api", "schema": "normalizado"})


@app.route("/api/auth/login", methods=["POST"])
@app.route("/api/login", methods=["POST"])
def login():
    data = parse_json()
    email = (data.get("email") or "").strip().lower()
    password = (data.get("password") or "").strip()
    if not email:
        return error("Correo requerido", 400)
    if not password:
        return error("Contraseña requerida", 400)

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM SOCIO WHERE email = %s", (email,))
            socio = cursor.fetchone()
            if not socio or socio["estado"] != "ACTIVO":
                return error("Credenciales inválidas", 401)

            valid = verify_password(password, socio.get("password_hash"))
            if not valid and socio.get("password_hash") in ("", None):
                valid = password == DEFAULT_PASSWORD
                if valid:
                    new_hash = hash_password(password)
                    cursor.execute(
                        "UPDATE SOCIO SET password_hash = %s, password_salt = '' WHERE id = %s",
                        (new_hash, socio["id"]),
                    )
                    socio["password_hash"] = new_hash

            if not valid:
                return error("Credenciales inválidas", 401)

            is_admin = socio.get("rol") == "ADMIN"
            conn.commit()

            token_payload = {
                "id": socio["id"],
                "email": socio["email"],
                "rol": "ADMIN" if is_admin else "SOCIO",
                "jti": str(uuid.uuid4()),
                "exp": datetime.utcnow() + timedelta(hours=JWT_EXPIRY_HOURS),
                "iat": datetime.utcnow(),
            }
            token = jwt.encode(token_payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

            return ok({"user": socio_public(socio), "is_admin": is_admin, "token": token})
    except Exception as exc:
        conn.rollback()
        logger.error("Login Error:", exc_info=True)
        return error("Error interno en el servidor al intentar iniciar sesión", 500)
    finally:
        conn.close()


@app.route("/api/auth/register", methods=["POST"])
def register():
    return create_user()


@app.route("/api/auth/logout", methods=["POST"])
@require_auth
def logout():
    auth_header = request.headers.get("Authorization", "")
    token = auth_header.split(" ", 1)[1] if " " in auth_header else ""
    if token:
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            jti = payload.get("jti")
            exp = payload.get("exp")
            if jti and exp:
                _revoke_token(jti, datetime.utcfromtimestamp(exp))
        except jwt.InvalidTokenError:
            pass
    return ok({"message": "Sesión cerrada correctamente"})


@app.route("/api/auth/recover", methods=["POST"])
def recover_password():
    data = parse_json()
    email = (data.get("email") or "").strip().lower()
    if not email:
        return error("Correo requerido", 400)

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, nombre, telefono FROM SOCIO WHERE email = %s AND estado = 'ACTIVO'", (email,))
            socio = cursor.fetchone()
            if not socio:
                # Respuesta genérica para evitar enumeración de emails
                return ok({"message": "Si el correo existe, recibirás tu PIN de acceso temporal."}, 200)
            
            pin = ''.join(secrets.choice(string.digits) for _ in range(6))
            new_hash = hash_password(pin)
            
            cursor.execute(
                "UPDATE SOCIO SET password_hash = %s, password_salt = '' WHERE id = %s",
                (new_hash, socio["id"])
            )
            conn.commit()
            
            primer_nombre = socio["nombre"].split(" ")[0]
            notify_whatsapp(
                socio["telefono"],
                f"Hola {primer_nombre}, se ha reseteado tu contraseña. Tu PIN temporal es: {pin}\nPor favor, ingresa al sistema y cámbiala en tu Perfil."
            )
            return ok({"message": "Se ha enviado un código temporal a tu WhatsApp"})
    except Exception as exc:
        conn.rollback()
        logger.error("Recover password error:", exc_info=True)
        return error("Error interno al intentar recuperar la contraseña", 500)
    finally:
        conn.close()


@app.route("/api/users", methods=["GET", "POST"])
@require_auth
def users_collection():
    if request.method == "POST":
        if request.user.get("rol") != "ADMIN":
            return error("No tienes permisos para crear usuarios", 403)
        return create_user()

    if request.user.get("rol") != "ADMIN":
        return error("Solo administradores pueden listar socios", 403)

    page = request.args.get("page", type=int)
    limit = request.args.get("limit", type=int)
    search = request.args.get("search", "").strip()

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            sql = "SELECT * FROM SOCIO WHERE 1=1"
            params = []
            if search:
                sql += " AND (nombre LIKE %s OR email LIKE %s)"
                params.extend([f"%{search}%", f"%{search}%"])
            
            cursor.execute(sql.replace("*", "COUNT(*) as total", 1), params)
            total = cursor.fetchone()["total"]
            
            sql += " ORDER BY nombre"
            if page and limit:
                offset = (page - 1) * limit
                sql += " LIMIT %s OFFSET %s"
                params.extend([limit, offset])
                
            cursor.execute(sql, params)
            data = [socio_public(row) for row in cursor.fetchall()]
            
            if page and limit:
                return ok({"data": data, "total": total, "page": page, "limit": limit})
            return ok(data)
    finally:
        conn.close()


def create_user():
    data = parse_json()
    required = require_fields(data, ["nombre_completo", "email_institucional", "telefono_whatsapp"])
    if required:
        return error(required)

    email = data["email_institucional"].strip().lower()
    if not _re.match(r'^[^@\s]+@pucesa\.edu\.ec$', email):
        return error("El correo debe tener el formato usuario@pucesa.edu.ec")

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO SOCIO (
                    nombre, email, telefono, password_hash, password_salt,
                    nivel_habilidad, rol, estado
                ) VALUES (%s, %s, %s, %s, '', %s, %s, %s)
                """,
                (
                    data["nombre_completo"].strip(),
                    email,
                    data["telefono_whatsapp"].strip(),
                    hash_password(data.get("password") or DEFAULT_PASSWORD),
                    normalize_level(data.get("nivel_habilidad")),
                    normalize_role(data.get("rol")),
                    normalize_socio_estado(data.get("estado")),
                ),
            )
            conn.commit()
            return ok({"message": "Socio agregado exitosamente", "id": cursor.lastrowid}, 201)
    except pymysql.err.IntegrityError:
        conn.rollback()
        return error("Ya existe un socio con ese correo", 409)
    except Exception as exc:
        conn.rollback()
        logger.error("Create User Error:", exc_info=True)
        return error("Error al crear el socio", 400)
    finally:
        conn.close()


@app.route('/api/socios/<int:socio_id>/historial', methods=['GET'])
@require_auth
def socio_historial(socio_id):
    # Only admin or the socio themselves can see their historial
    if request.user.get('rol') != 'ADMIN' and request.user.get('id') != socio_id:
        return error('No autorizado', 403)
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT 'reserva' AS tipo, r.id, r.fecha_inicio AS fecha, r.estado,
                       sa.nombre AS detalle
                FROM RESERVA r
                JOIN SALA sa ON sa.id = r.sala_id
                WHERE r.socio_id = %s
                UNION ALL
                SELECT 'prestamo' AS tipo, p.id, p.fecha_salida AS fecha, p.estado,
                       i.nombre AS detalle
                FROM PRESTAMO p
                JOIN INSTRUMENTO i ON i.id = p.instrumento_id
                WHERE p.socio_id = %s
                ORDER BY fecha DESC
                LIMIT 50
            """, (socio_id, socio_id))
            items = serialize_rows(cursor.fetchall())
            return ok(items)
    finally:
        conn.close()


@app.route("/api/users/<int:user_id>", methods=["GET", "PUT", "DELETE"])
@require_auth
def user_detail(user_id):
    is_admin = request.user.get("rol") == "ADMIN"
    requester_id = request.user.get("id")

    if request.method == "GET":
        if not is_admin and requester_id != user_id:
            return error("No autorizado", 403)
    elif request.method in ("PUT", "DELETE"):
        if not is_admin:
            return error("Solo administradores pueden modificar socios", 403)

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            if request.method == "GET":
                cursor.execute("SELECT * FROM SOCIO WHERE id = %s", (user_id,))
                socio = cursor.fetchone()
                return ok(socio_public(socio)) if socio else error("Socio no encontrado", 404)

            if request.method == "DELETE":
                # An admin cannot inactivate themselves
                if user_id == requester_id:
                    return error(
                        "No puedes inactivar tu propia cuenta. Pídele a otro administrador que lo haga.",
                        403,
                    )
                # Prevent leaving the system with zero active admins
                cursor.execute("SELECT rol FROM SOCIO WHERE id = %s", (user_id,))
                target = cursor.fetchone()
                if target and target.get("rol") == "ADMIN":
                    cursor.execute(
                        "SELECT COUNT(*) AS total FROM SOCIO WHERE rol = 'ADMIN' AND estado = 'ACTIVO' AND id != %s",
                        (user_id,),
                    )
                    if cursor.fetchone()["total"] == 0:
                        return error(
                            "No puedes inactivar al único administrador activo. "
                            "Asigna primero otro administrador.",
                            403,
                        )
                cursor.execute("UPDATE SOCIO SET estado = 'INACTIVO' WHERE id = %s", (user_id,))
                conn.commit()
                return ok({"message": "Socio inactivado"})

            data = parse_json()

            # An admin can never change their own rol — only another admin can do it
            if "rol" in data and user_id == requester_id:
                return error(
                    "No puedes cambiar tu propio rol. Pídele a otro administrador que lo haga.",
                    403,
                )

            # Prevent leaving the system with zero active admins
            if "rol" in data:
                new_rol = normalize_role(data.get("rol"))
                if new_rol != "ADMIN":
                    cursor.execute(
                        "SELECT rol FROM SOCIO WHERE id = %s", (user_id,)
                    )
                    target_socio = cursor.fetchone()
                    if target_socio and target_socio.get("rol") == "ADMIN":
                        cursor.execute(
                            "SELECT COUNT(*) AS total FROM SOCIO WHERE rol = 'ADMIN' AND estado = 'ACTIVO' AND id != %s",
                            (user_id,),
                        )
                        if cursor.fetchone()["total"] == 0:
                            return error(
                                "No puedes degradar al único administrador activo. "
                                "Asigna primero otro administrador.",
                                403,
                            )

            mapping = {
                "nombre_completo": ("nombre", lambda value: value.strip()),
                "telefono_whatsapp": ("telefono", lambda value: value.strip()),
                "nivel_habilidad": ("nivel_habilidad", normalize_level),
                "rol": ("rol", normalize_role),
                "estado": ("estado", normalize_socio_estado),
            }
            updates = [(column, transform(data[field])) for field, (column, transform) in mapping.items() if field in data]
            if not updates:
                return error("No hay campos para actualizar")
            _SOCIO_COLS = {"nombre", "telefono", "nivel_habilidad", "rol", "estado"}
            sql = _build_safe_set(updates, _SOCIO_COLS)
            cursor.execute(f"UPDATE SOCIO SET {sql} WHERE id = %s", [value for _, value in updates] + [user_id])  # nosec B608
            conn.commit()
            return ok({"message": "Socio actualizado"})
    except Exception as exc:
        conn.rollback()
        logger.error("Internal error", exc_info=True)
        return error("Error interno del servidor", 500)
    finally:
        conn.close()


@app.route('/api/users/me/avatar', methods=['PUT'])
@require_auth
def update_avatar():
    data = parse_json()
    avatar_url = data.get('avatar_url', '').strip()
    if not avatar_url:
        return error('avatar_url requerido')
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute('UPDATE SOCIO SET avatar_url = %s WHERE id = %s',
                           (avatar_url, request.user['id']))
            conn.commit()
            return ok({'message': 'Avatar actualizado'})
    except Exception as exc:
        conn.rollback()
        logger.error("Internal error", exc_info=True)
        return error("Error interno del servidor", 500)
    finally:
        conn.close()


@app.route("/api/users/me", methods=["PUT"])
@require_auth
def update_my_profile():
    data = parse_json()
    updates = []
    if "telefono_whatsapp" in data:
        updates.append(("telefono", data["telefono_whatsapp"].strip()))
    if "nivel_habilidad" in data:
        updates.append(("nivel_habilidad", normalize_level(data["nivel_habilidad"])))
    
    if not updates:
        return error("No hay campos válidos para actualizar")

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            _PERFIL_COLS = {"telefono", "nivel_habilidad"}
            sql = _build_safe_set(updates, _PERFIL_COLS)
            cursor.execute(
                f"UPDATE SOCIO SET {sql} WHERE id = %s",  # nosec B608
                [value for _, value in updates] + [request.user["id"]]
            )
            conn.commit()
            return ok({"message": "Perfil actualizado correctamente"})
    except Exception as exc:
        conn.rollback()
        logger.error("Update profile error:", exc_info=True)
        return error("Error interno al actualizar perfil", 500)
    finally:
        conn.close()


@app.route("/api/users/me/password", methods=["PUT"])
@require_auth
def update_my_password():
    data = parse_json()
    current_password = data.get("current_password")
    new_password = data.get("new_password")
    
    if not current_password or not new_password:
        return error("Se requiere contraseña actual y nueva", 400)

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT password_hash FROM SOCIO WHERE id = %s", (request.user["id"],))
            socio = cursor.fetchone()
            
            valid = verify_password(current_password, socio.get("password_hash"))
            if not valid and socio.get("password_hash") in ("", None):
                valid = current_password == DEFAULT_PASSWORD
                
            if not valid:
                return error("La contraseña actual es incorrecta", 401)
                
            new_hash = hash_password(new_password)
            cursor.execute(
                "UPDATE SOCIO SET password_hash = %s, password_salt = '' WHERE id = %s",
                (new_hash, request.user["id"])
            )
            conn.commit()
            return ok({"message": "Contraseña actualizada correctamente"})
    except Exception as exc:
        conn.rollback()
        logger.error("Update password error:", exc_info=True)
        return error("Error interno al actualizar contraseña", 500)
    finally:
        conn.close()


@app.route("/api/inventario", methods=["GET", "POST"])
@app.route("/api/instrumentos", methods=["GET", "POST"])
@require_auth
def inventory_collection():
    if request.method == "POST":
        if request.user.get("rol") != "ADMIN":
            return error("No tienes permisos para agregar instrumentos", 403)
        return create_inventory_item()

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT i.id, i.nombre, i.tipo_instrumento_id, ti.nombre AS tipo,
                       i.marca, i.modelo, i.numero_serie, i.fecha_adquisicion,
                       i.estado, i.ubicacion, i.ultima_actualizacion,
                       CASE WHEN i.estado = 'DISPONIBLE' AND p.id IS NULL THEN 1 ELSE 0 END AS disponible
                FROM INSTRUMENTO i
                JOIN TIPO_INSTRUMENTO ti ON ti.id = i.tipo_instrumento_id
                LEFT JOIN PRESTAMO p ON p.instrumento_id = i.id
                    AND p.estado = 'ACTIVO'
                    AND p.eliminado_en IS NULL
                ORDER BY ti.nombre, i.nombre
                """
            )
            return ok(serialize_rows(cursor.fetchall()))
    finally:
        conn.close()


@app.route("/api/instrumentos/disponibles", methods=["GET"])
@require_auth
def available_instruments():
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT i.id, i.nombre, i.tipo_instrumento_id, ti.nombre AS tipo,
                       i.marca, i.modelo, i.numero_serie, i.estado, i.ubicacion
                FROM INSTRUMENTO i
                JOIN TIPO_INSTRUMENTO ti ON ti.id = i.tipo_instrumento_id
                LEFT JOIN PRESTAMO p ON p.instrumento_id = i.id
                    AND p.estado = 'ACTIVO'
                    AND p.eliminado_en IS NULL
                WHERE i.estado = 'DISPONIBLE' AND p.id IS NULL
                ORDER BY i.nombre
                """
            )
            return ok(serialize_rows(cursor.fetchall()))
    finally:
        conn.close()


def create_inventory_item():
    data = parse_json()
    required = require_fields(data, ["nombre", "tipo", "ubicacion"])
    if required:
        return error(required)

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            tipo_id = get_or_create_tipo_instrumento(cursor, data.get("tipo"))
            numero_serie = (data.get("numero_serie") or f"AUTO-{datetime.now().strftime('%Y%m%d%H%M%S%f')}").strip()
            cursor.execute(
                """
                INSERT INTO INSTRUMENTO (
                    nombre, tipo_instrumento_id, marca, modelo, numero_serie,
                    fecha_adquisicion, estado, ubicacion, modificado_por
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    data["nombre"].strip(),
                    tipo_id,
                    data.get("marca") or None,
                    data.get("modelo") or None,
                    numero_serie,
                    data.get("fecha_adquisicion") or None,
                    normalize_instrumento_estado(data.get("estado")),
                    data["ubicacion"].strip(),
                    data.get("modificado_por", "API"),
                ),
            )
            conn.commit()
            return ok({"message": "Instrumento agregado exitosamente", "id": cursor.lastrowid}, 201)
    except Exception as exc:
        conn.rollback()
        logger.error("Create Inventory Error:", exc_info=True)
        return error("Error al agregar el instrumento", 400)
    finally:
        conn.close()


@app.route("/api/inventario/<int:item_id>", methods=["GET", "PUT"])
@app.route("/api/instrumentos/<int:item_id>", methods=["GET", "PUT"])
@require_auth
def inventory_detail(item_id):
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            if request.method == "GET":
                cursor.execute(
                    """
                    SELECT i.*, ti.nombre AS tipo
                    FROM INSTRUMENTO i
                    JOIN TIPO_INSTRUMENTO ti ON ti.id = i.tipo_instrumento_id
                    WHERE i.id = %s
                    """,
                    (item_id,),
                )
                item = cursor.fetchone()
                return ok(serialize_row(item)) if item else error("Instrumento no encontrado", 404)

            if request.user.get("rol") != "ADMIN":
                return error("No tienes permisos para modificar instrumentos", 403)
            data = parse_json()

            # Leer estado actual del instrumento
            cursor.execute("SELECT estado FROM INSTRUMENTO WHERE id = %s", (item_id,))
            current_inst = cursor.fetchone()
            if not current_inst:
                return error("Instrumento no encontrado", 404)
            current_estado = current_inst["estado"]

            updates = []
            if "tipo" in data:
                updates.append(("tipo_instrumento_id", get_or_create_tipo_instrumento(cursor, data.get("tipo"))))
            for field in ["nombre", "marca", "modelo", "numero_serie", "ubicacion"]:
                if field in data:
                    updates.append((field, data[field]))

            if "estado" in data:
                nuevo_estado = normalize_instrumento_estado(data.get("estado"))

                # Bloquear cambios manuales si el instrumento está PRESTADO
                if current_estado == "PRESTADO":
                    return error(
                        "No se puede cambiar el estado de un instrumento que está prestado. "
                        "Primero registra la devolución del préstamo activo.",
                        409,
                    )

                # Transiciones válidas
                _VALID_TRANSITIONS = {
                    "DISPONIBLE":    {"MANTENIMIENTO", "BAJA"},
                    "MANTENIMIENTO": {"DISPONIBLE", "BAJA"},
                    "BAJA":          {"DISPONIBLE", "MANTENIMIENTO"},
                    "PRESTADO":      set(),  # solo el sistema de préstamos puede cambiar este
                }
                allowed = _VALID_TRANSITIONS.get(current_estado, set())
                if nuevo_estado != current_estado and nuevo_estado not in allowed:
                    return error(
                        f"Transición de estado no permitida: {current_estado} → {nuevo_estado}",
                        400,
                    )

                # Verificar doble seguridad: no debe haber préstamo activo aunque el estado no diga PRESTADO
                cursor.execute(
                    "SELECT id FROM PRESTAMO WHERE instrumento_id = %s AND estado = 'ACTIVO' AND eliminado_en IS NULL",
                    (item_id,),
                )
                if cursor.fetchone():
                    return error(
                        "Este instrumento tiene un préstamo activo registrado. "
                        "Registra la devolución antes de modificar su estado.",
                        409,
                    )

                updates.append(("estado", nuevo_estado))

            updates.append(("modificado_por", data.get("modificado_por", "API")))
            _INSTRUMENTO_COLS = {"nombre", "tipo_instrumento_id", "marca", "modelo", "numero_serie",
                                  "estado", "ubicacion", "modificado_por"}
            sql = _build_safe_set(updates, _INSTRUMENTO_COLS)
            cursor.execute(f"UPDATE INSTRUMENTO SET {sql} WHERE id = %s", [value for _, value in updates] + [item_id])  # nosec B608
            conn.commit()
            return ok({"message": "Instrumento actualizado"})
    except Exception as exc:
        conn.rollback()
        logger.error("Update Inventory Error:", exc_info=True)
        return error("Error al actualizar el instrumento", 400)
    finally:
        conn.close()


@app.route("/api/salas", methods=["GET", "POST"])
@require_auth
def rooms_collection():
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            if request.method == "POST":
                if request.user.get("rol") != "ADMIN":
                    return error("No tienes permisos para crear salas", 403)
                data = parse_json()
                required = require_fields(data, ["nombre", "capacidad"])
                if required:
                    return error(required)
                cursor.execute(
                    """
                    INSERT INTO SALA (nombre, tipo, capacidad, equipamiento, estado)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (
                        data["nombre"].strip(),
                        normalize_sala_tipo(data.get("tipo")),
                        int(data["capacidad"]),
                        data.get("equipamiento"),
                        normalize_sala_estado(data.get("estado")),
                    ),
                )
                conn.commit()
                return ok({"message": "Sala agregada", "id": cursor.lastrowid}, 201)

            cursor.execute("SELECT * FROM SALA ORDER BY nombre")
            return ok(serialize_rows(cursor.fetchall()))
    except Exception as exc:
        conn.rollback()
        logger.error("Salas Error:", exc_info=True)
        return error("Error interno al procesar la sala", 400)
    finally:
        conn.close()


@app.route("/api/salas/<int:sala_id>", methods=["PUT"])
@require_admin
def room_detail(sala_id):
    data = parse_json()
    updates = []
    for field in ["nombre", "capacidad", "equipamiento"]:
        if field in data:
            updates.append((field, data[field]))
    if "tipo" in data:
        updates.append(("tipo", normalize_sala_tipo(data.get("tipo"))))
    if "estado" in data:
        updates.append(("estado", normalize_sala_estado(data.get("estado"))))
    if not updates:
        return error("No hay campos para actualizar")

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            _SALA_COLS = {"nombre", "capacidad", "equipamiento", "tipo", "estado"}
            sql = _build_safe_set(updates, _SALA_COLS)
            cursor.execute(f"UPDATE SALA SET {sql} WHERE id = %s", [value for _, value in updates] + [sala_id])  # nosec B608
            conn.commit()
            return ok({"message": "Sala actualizada"})
    except Exception as exc:
        conn.rollback()
        logger.error("Room Update Error:", exc_info=True)
        return error("Error al actualizar la sala", 400)
    finally:
        conn.close()


@app.route("/api/reservas", methods=["GET", "POST"])
@require_auth
def reservations_collection():
    if request.method == "POST":
        return create_reservation()

    page = request.args.get("page", type=int)
    limit = request.args.get("limit", type=int)
    search = request.args.get("search", "").strip()
    
    is_admin = request.user.get("rol") == "ADMIN"
    user_id = request.user.get("id")

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            sql = """
                SELECT r.id, r.socio_id AS user_id, r.sala_id, r.fecha_inicio, r.fecha_fin,
                       r.estado, r.observaciones, r.fecha_creacion,
                       s.nombre AS nombre_completo, sa.nombre AS sala_nombre
                FROM RESERVA r
                JOIN SOCIO s ON s.id = r.socio_id
                JOIN SALA sa ON sa.id = r.sala_id
                WHERE r.eliminado_en IS NULL
            """
            params = []
            
            if not is_admin:
                sql += " AND r.socio_id = %s"
                params.append(user_id)
            
            if search:
                sql += " AND (s.nombre LIKE %s OR sa.nombre LIKE %s)"
                params.extend([f"%{search}%", f"%{search}%"])
                
            count_sql = "SELECT COUNT(*) as total FROM RESERVA r JOIN SOCIO s ON s.id = r.socio_id JOIN SALA sa ON sa.id = r.sala_id WHERE r.eliminado_en IS NULL"
            count_params = []
            if not is_admin:
                count_sql += " AND r.socio_id = %s"
                count_params.append(user_id)
                
            if search:
                count_sql += " AND (s.nombre LIKE %s OR sa.nombre LIKE %s)"
                count_params.extend([f"%{search}%", f"%{search}%"])
            
            cursor.execute(count_sql, count_params)
            total = cursor.fetchone()["total"]

            sql += " ORDER BY r.fecha_inicio DESC"
            if page and limit:
                offset = (page - 1) * limit
                sql += " LIMIT %s OFFSET %s"
                params.extend([limit, offset])

            cursor.execute(sql, params)
            data = serialize_rows(cursor.fetchall())
            if page and limit:
                return ok({"data": data, "total": total, "page": page, "limit": limit})
            return ok(data)
    finally:
        conn.close()


def check_reservation_rules(fecha_inicio, fecha_fin):
    ahora_local = datetime.utcnow() - timedelta(hours=5)
    if fecha_inicio < ahora_local:
        return "La fecha de inicio no puede estar en el pasado"
    minutos_anticipacion = (fecha_inicio - ahora_local).total_seconds() / 60
    if minutos_anticipacion < 30:
        return 'Debes reservar con al menos 30 minutos de anticipación'
    if fecha_fin <= fecha_inicio:
        return "La fecha de fin debe ser posterior al inicio"
    if fecha_inicio.date() != fecha_fin.date():
        return "La reserva debe iniciar y terminar el mismo día"

    if fecha_inicio.hour < 8 or fecha_fin.hour > 22 or (fecha_fin.hour == 22 and fecha_fin.minute > 0):
        return "Las reservas solo están permitidas entre las 08:00 y las 22:00"

    duracion = (fecha_fin - fecha_inicio).total_seconds() / 3600
    if duracion < 1:
        return "La reserva debe durar al menos 1 hora"
    if duracion > 4:
        return "La reserva no puede durar más de 4 horas"
    return None


@app.route("/api/reservas/validar", methods=["POST"])
@require_auth
def validate_reservation():
    data = parse_json()
    required = require_fields(data, ["sala_id", "fecha_inicio", "fecha_fin"])
    if required:
        return error(required)
    try:
        fecha_inicio = parse_datetime(data["fecha_inicio"], "fecha_inicio")
        fecha_fin = parse_datetime(data["fecha_fin"], "fecha_fin")
    except ValueError as exc:
        return error(str(exc))
    
    rule_error = check_reservation_rules(fecha_inicio, fecha_fin)
    if rule_error:
        return error(rule_error)

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            available = not reservation_overlap(cursor, data["sala_id"], fecha_inicio, fecha_fin)
            return ok({"disponible": available})
    finally:
        conn.close()


def create_reservation():
    if not within_operating_hours():
        return error(OPERATING_HOURS_MSG, 403)
    data = parse_json()
    required = require_fields(data, ["user_id", "sala_id", "fecha_inicio", "fecha_fin"])
    if required:
        return error(required)

    # Un SOCIO solo puede crear reservas para sí mismo
    is_admin = request.user.get("rol") == "ADMIN"
    requester_id = request.user.get("id")
    target_user_id = int(data["user_id"])
    if not is_admin and target_user_id != requester_id:
        return error("No puedes crear reservas para otro usuario", 403)

    try:
        fecha_inicio = parse_datetime(data["fecha_inicio"], "fecha_inicio")
        fecha_fin = parse_datetime(data["fecha_fin"], "fecha_fin")
    except ValueError as exc:
        return error(str(exc))

    if fecha_inicio < _now_local():
        return error("La fecha de inicio no puede estar en el pasado", 400)

    if not is_valid_operating_date(fecha_inicio):
        return error(OPERATING_DATE_MSG, 400)

    rule_error = check_reservation_rules(fecha_inicio, fecha_fin)
    if rule_error:
        return error(rule_error)

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute('SELECT nombre, estado FROM SALA WHERE id = %s', (data['sala_id'],))
            sala = cursor.fetchone()
            if not sala:
                return error("Sala no encontrada", 404)
            if sala["estado"] != "ACTIVA":
                return error("La sala no está activa")
            if reservation_overlap(cursor, data["sala_id"], fecha_inicio, fecha_fin):
                return error("La sala ya tiene una reserva en ese horario", 409)

            cursor.execute(
                """
                INSERT INTO RESERVA (
                    socio_id, sala_id, fecha_inicio, fecha_fin, estado,
                    observaciones, creado_por, modificado_por
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    target_user_id,
                    data["sala_id"],
                    fecha_inicio,
                    fecha_fin,
                    normalize_reserva_estado(data.get("estado")),
                    data.get("observaciones"),
                    data.get("creado_por", "API"),
                    data.get("modificado_por", "API"),
                ),
            )
            reserva_id = cursor.lastrowid
            cursor.execute('SELECT nombre, telefono, email FROM SOCIO WHERE id = %s', (data['user_id'],))
            socio = cursor.fetchone()
            conn.commit()

            notified = False
            if socio:
                primer_nombre = socio['nombre'].split(' ')[0]
                notified = notify_whatsapp(
                    socio['telefono'],
                    f"\U0001f3b5 Hola {primer_nombre}! Tu reserva de *{sala['nombre']}* fue confirmada.\n\U0001f4c5 Inicio: {fecha_inicio.strftime('%d/%m/%Y %H:%M')}\n\U0001f514 Fin: {fecha_fin.strftime('%d/%m/%Y %H:%M')}\nID: #{reserva_id}",
                )
                if socio.get('email'):
                    html = build_email_reserva_confirmada(
                        socio['nombre'], sala['nombre'],
                        fecha_inicio.strftime('%A %d de %B %Y, %H:%M'),
                        fecha_fin.strftime('%H:%M'),
                        reserva_id
                    )
                    send_email_async(socio['email'], '\u2705 Reserva Confirmada \u2014 Club de M\u00fasica', html)
            return ok({'message': 'Reserva creada exitosamente', 'id': reserva_id, 'whatsapp_enviado': notified}, 201)
    except Exception as exc:
        conn.rollback()
        logger.error("Create Reservation Error:", exc_info=True)
        return error("Error interno al crear la reserva", 400)
    finally:
        conn.close()


@app.route("/api/reservas/<int:user_id>", methods=["GET"])
@require_auth
def get_reservas_usuario(user_id):
    if request.user.get("rol") != "ADMIN" and request.user.get("id") != user_id:
        return error("No autorizado", 403)
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT r.id, r.socio_id AS user_id, r.sala_id, r.fecha_inicio, r.fecha_fin,
                       r.estado, r.observaciones, sa.nombre AS sala_nombre
                FROM RESERVA r
                JOIN SALA sa ON sa.id = r.sala_id
                WHERE r.socio_id = %s AND r.eliminado_en IS NULL
                ORDER BY r.fecha_inicio DESC
                """,
                (user_id,),
            )
            return ok(serialize_rows(cursor.fetchall()))
    finally:
        conn.close()


@app.route("/api/reservas/calendario", methods=["GET"])
@require_auth
def reservation_calendar():
    start = request.args.get("start")
    end = request.args.get("end")
    is_admin = request.user.get("rol") == "ADMIN"
    user_id = request.user.get("id")

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            sql = """
                SELECT r.id, r.socio_id, r.sala_id, r.fecha_inicio AS start, r.fecha_fin AS end, r.estado,
                       s.nombre AS nombre_completo, sa.nombre AS sala_nombre
                FROM RESERVA r
                JOIN SOCIO s ON s.id = r.socio_id
                JOIN SALA sa ON sa.id = r.sala_id
                WHERE r.eliminado_en IS NULL
                  AND r.estado IN ('CONFIRMADA', 'COMPLETADA', 'REPROGRAMADA')
            """
            params = []
            if start:
                sql += " AND r.fecha_inicio >= %s"
                params.append(start)
            if end:
                sql += " AND r.fecha_fin <= %s"
                params.append(end)
            
            sql += " ORDER BY r.fecha_inicio"
            cursor.execute(sql, params)
            
            events = []
            for row in serialize_rows(cursor.fetchall()):
                if is_admin or row["socio_id"] == user_id:
                    row["title"] = f"{row['sala_nombre']} - {row['nombre_completo']}"
                else:
                    row["title"] = f"{row['sala_nombre']} - Reservada"
                    row["nombre_completo"] = "Ocupado"
                events.append(row)
            return ok(events)
    finally:
        conn.close()


@app.route('/api/reservas/<int:reserva_id>', methods=['DELETE'])
@require_auth
def cancel_reservation(reserva_id):
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            # Get reservation + socio details before canceling
            cursor.execute("""
                SELECT r.id, r.socio_id, r.fecha_inicio, r.fecha_fin, r.sala_id,
                       s.nombre, s.telefono, s.email, sa.nombre AS sala_nombre
                FROM RESERVA r
                JOIN SOCIO s ON s.id = r.socio_id
                JOIN SALA sa ON sa.id = r.sala_id
                WHERE r.id = %s AND r.eliminado_en IS NULL
            """, (reserva_id,))
            reserva = cursor.fetchone()

            if reserva:
                is_admin = request.user.get("rol") == "ADMIN"
                is_owner = request.user.get("id") == reserva["socio_id"]
                if not is_admin and not is_owner:
                    return error("No autorizado para cancelar esta reserva", 403)

            cursor.execute("""
                UPDATE RESERVA
                SET estado = 'CANCELADA', eliminado_en = NOW(), modificado_por = 'API'
                WHERE id = %s AND eliminado_en IS NULL
            """, (reserva_id,))
            conn.commit()

            if reserva:
                primer_nombre = reserva['nombre'].split(' ')[0]
                fecha_str = reserva['fecha_inicio'].strftime('%d/%m/%Y %H:%M') if isinstance(reserva['fecha_inicio'], datetime) else str(reserva['fecha_inicio'])
                notify_whatsapp(
                    reserva['telefono'],
                    f"\u274c Hola {primer_nombre}, tu reserva de *{reserva['sala_nombre']}* del {fecha_str} ha sido CANCELADA. Puedes hacer una nueva reserva en el sistema."
                )
                if reserva.get('email'):
                    html = build_email_reserva_cancelada(
                        reserva['nombre'], reserva['sala_nombre'], fecha_str, reserva_id
                    )
                    send_email_async(reserva['email'], '\u274c Reserva Cancelada \u2014 Club de M\u00fasica', html)

            return ok({'message': 'Reserva cancelada'})
    except Exception as exc:
        conn.rollback()
        logger.error('Cancel Reservation Error:', exc_info=True)
        return error('Error al cancelar la reserva', 400)
    finally:
        conn.close()


@app.route("/api/prestamos", methods=["GET"])
@require_auth
def loans_collection():
    is_admin = request.user.get("rol") == "ADMIN"
    requester_id = request.user.get("id")
    estado = request.args.get("estado")
    page = request.args.get("page", type=int)
    limit = request.args.get("limit", type=int)
    search = request.args.get("search", "").strip()
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            sql = """
                SELECT p.id, p.socio_id AS user_id, p.instrumento_id AS inventario_id,
                       p.fecha_salida, p.fecha_limite, p.fecha_devolucion, p.estado,
                       p.motivo, p.observaciones, s.nombre AS nombre_completo,
                       i.nombre AS instrumento_nombre, ti.nombre AS tipo
                FROM PRESTAMO p
                JOIN SOCIO s ON s.id = p.socio_id
                JOIN INSTRUMENTO i ON i.id = p.instrumento_id
                JOIN TIPO_INSTRUMENTO ti ON ti.id = i.tipo_instrumento_id
                WHERE p.eliminado_en IS NULL
            """
            params = []
            if not is_admin:
                sql += " AND p.socio_id = %s"
                params.append(requester_id)
            if estado:
                sql += " AND p.estado = %s"
                params.append(normalize_prestamo_estado(estado))

            if search and is_admin:
                sql += " AND (s.nombre LIKE %s OR i.nombre LIKE %s OR i.numero_serie LIKE %s)"
                params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])

            # Count total
            count_sql = "SELECT COUNT(*) as total FROM PRESTAMO p JOIN SOCIO s ON s.id = p.socio_id JOIN INSTRUMENTO i ON i.id = p.instrumento_id JOIN TIPO_INSTRUMENTO ti ON ti.id = i.tipo_instrumento_id WHERE p.eliminado_en IS NULL"
            count_params = []
            if not is_admin:
                count_sql += " AND p.socio_id = %s"
                count_params.append(requester_id)
            if estado:
                count_sql += " AND p.estado = %s"
                count_params.append(normalize_prestamo_estado(estado))
            if search and is_admin:
                count_sql += " AND (s.nombre LIKE %s OR i.nombre LIKE %s OR i.numero_serie LIKE %s)"
                count_params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
            cursor.execute(count_sql, count_params)
            total = cursor.fetchone()["total"]

            sql += " ORDER BY p.fecha_salida DESC"
            if page and limit:
                offset = (page - 1) * limit
                sql += " LIMIT %s OFFSET %s"
                params.extend([limit, offset])
                
            cursor.execute(sql, params)
            data = serialize_rows(cursor.fetchall())
            
            if page and limit:
                return ok({"data": data, "total": total, "page": page, "limit": limit})
            return ok(data)
    finally:
        conn.close()


@app.route("/api/prestamos/activos", methods=["GET"])
@require_auth
def active_loans():
    return loans_by_status("ACTIVO")


def loans_by_status(status):
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT p.id, p.socio_id AS user_id, p.instrumento_id AS inventario_id,
                       p.fecha_salida, p.fecha_limite, p.fecha_devolucion, p.estado,
                       p.motivo, p.observaciones, s.nombre AS nombre_completo,
                       i.nombre AS instrumento_nombre, ti.nombre AS tipo
                FROM PRESTAMO p
                JOIN SOCIO s ON s.id = p.socio_id
                JOIN INSTRUMENTO i ON i.id = p.instrumento_id
                JOIN TIPO_INSTRUMENTO ti ON ti.id = i.tipo_instrumento_id
                WHERE p.estado = %s AND p.eliminado_en IS NULL
                ORDER BY p.fecha_limite
                """,
                (status,),
            )
            return ok(serialize_rows(cursor.fetchall()))
    finally:
        conn.close()


@app.route("/api/prestamos/solicitar", methods=["POST"])
@require_auth
def request_loan():
    if not within_operating_hours():
        return error(OPERATING_HOURS_MSG, 403)
    data = parse_json()
    required = require_fields(data, ["user_id", "inventario_id", "fecha_salida", "fecha_limite", "motivo"])
    if required:
        return error(required)

    # Un SOCIO solo puede pedir préstamos para sí mismo
    _is_admin = request.user.get("rol") == "ADMIN"
    _requester_id = request.user.get("id")
    if not _is_admin and int(data["user_id"]) != _requester_id:
        return error("No puedes solicitar préstamos en nombre de otro usuario", 403)

    try:
        fecha_salida = parse_datetime(data["fecha_salida"], "fecha_salida")
        fecha_limite = parse_datetime(data["fecha_limite"], "fecha_limite")
    except ValueError as exc:
        return error(str(exc))

    if not is_valid_operating_date(fecha_salida):
        return error(OPERATING_DATE_MSG, 400)

    if not is_valid_operating_date(fecha_limite):
        return error(
            "La fecha límite de devolución no puede ser domingo ni sábado a partir de las 12:00.",
            400,
        )

    if fecha_limite <= fecha_salida:
        return error("La fecha y hora límite debe ser posterior a la de salida", 400)

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, estado, nombre AS instrumento_nombre
                FROM INSTRUMENTO
                WHERE id = %s
                FOR UPDATE
                """,
                (data["inventario_id"],),
            )
            instrument = cursor.fetchone()
            if not instrument:
                return error("Instrumento no encontrado", 404)
            if instrument["estado"] != "DISPONIBLE":
                return error("El instrumento no está disponible", 409)

            cursor.execute(
                """
                SELECT id FROM PRESTAMO
                WHERE instrumento_id = %s AND estado = 'ACTIVO' AND eliminado_en IS NULL
                """,
                (data["inventario_id"],),
            )
            if cursor.fetchone():
                return error("El instrumento ya está prestado", 409)

            # Un socio solo puede tener un préstamo activo a la vez
            cursor.execute(
                """
                SELECT COUNT(*) AS total FROM PRESTAMO
                WHERE socio_id = %s AND estado = 'ACTIVO' AND eliminado_en IS NULL
                """,
                (data["user_id"],),
            )
            if cursor.fetchone()["total"] > 0:
                return error(
                    "Este socio ya tiene un préstamo activo. "
                    "Debe devolver el instrumento actual antes de solicitar otro.",
                    409,
                )

            cursor.execute(
                "SELECT nombre, telefono FROM SOCIO WHERE id = %s",
                (data["user_id"],),
            )
            socio_info = cursor.fetchone()

            observaciones = data.get("observaciones") or data.get("evento_universidad") or ""
            documento = data.get("documento_garantia")
            if documento:
                observaciones = f"{observaciones} | Garantía: {documento}".strip(" |")
            cursor.execute(
                """
                INSERT INTO PRESTAMO (
                    socio_id, instrumento_id, fecha_salida, fecha_limite, motivo, estado,
                    observaciones, creado_por, modificado_por
                ) VALUES (%s, %s, %s, %s, %s, 'ACTIVO', %s, %s, %s)
                """,
                (
                    data["user_id"],
                    data["inventario_id"],
                    fecha_salida,
                    fecha_limite,
                    data["motivo"],
                    observaciones,
                    data.get("creado_por", "API"),
                    data.get("modificado_por", "API"),
                ),
            )
            loan_id = cursor.lastrowid
            cursor.execute(
                "UPDATE INSTRUMENTO SET estado = 'PRESTADO', modificado_por = 'API' WHERE id = %s",
                (data["inventario_id"],),
            )
            conn.commit()

            # Notify socio via WhatsApp (non-fatal)
            try:
                if socio_info and socio_info.get("telefono"):
                    primer_nombre = socio_info["nombre"].split(" ")[0]
                    fecha_str = fecha_limite.strftime("%d/%m/%Y")
                    notify_whatsapp(
                        socio_info["telefono"],
                        f"✅ Hola {primer_nombre}, se registró tu préstamo de *{instrument['instrumento_nombre']}* hasta el {fecha_str}. ¡Cuídalo bien y devúelvelo a tiempo!",
                    )
            except Exception:
                logger.warning("WhatsApp notification failed for loan %s", loan_id, exc_info=True)

            return ok({"message": "Préstamo registrado", "id": loan_id}, 201)
    except Exception as exc:
        conn.rollback()
        logger.error("Request Loan Error:", exc_info=True)
        return error("Error interno al solicitar el préstamo", 400)
    finally:
        conn.close()


@app.route("/api/prestamos/<int:prestamo_id>/devolver", methods=["POST"])
@require_auth
def return_loan(prestamo_id):
    if not within_operating_hours():
        return error(OPERATING_HOURS_MSG, 403)
    if request.user.get("rol") != "ADMIN":
        return error("Solo administradores pueden registrar devoluciones", 403)
    data = parse_json()
    estado_instrumento = normalize_instrumento_estado(data.get("estado_instrumento", "DISPONIBLE"))
    if estado_instrumento == "PRESTADO":
        estado_instrumento = "DISPONIBLE"

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT instrumento_id, estado FROM PRESTAMO WHERE id = %s AND eliminado_en IS NULL",
                (prestamo_id,),
            )
            loan = cursor.fetchone()
            if not loan:
                return error("Préstamo no encontrado", 404)
            if loan["estado"] == "DEVUELTO":
                return error("Este préstamo ya fue devuelto anteriormente", 409)
            cursor.execute(
                """
                UPDATE PRESTAMO
                SET estado = 'DEVUELTO', fecha_devolucion = CURRENT_TIMESTAMP, modificado_por = 'API'
                WHERE id = %s AND estado IN ('ACTIVO', 'VENCIDO')
                """,
                (prestamo_id,),
            )
            cursor.execute(
                "UPDATE INSTRUMENTO SET estado = %s, modificado_por = 'API' WHERE id = %s",
                (estado_instrumento, loan["instrumento_id"]),
            )
            conn.commit()
            return ok({"message": "Préstamo devuelto"})
    except Exception as exc:
        conn.rollback()
        logger.error("Return Loan Error:", exc_info=True)
        return error("Error al registrar la devolución", 400)
    finally:
        conn.close()


@app.route("/api/dashboard/stats", methods=["GET"])
@require_auth
def dashboard_stats():
    conn = get_db_connection()
    is_admin = request.user.get("rol") == "ADMIN"
    user_id = request.user.get("id")
    
    try:
        with conn.cursor() as cursor:
            stats = {}
            if is_admin:
                queries = {
                    "socios_activos": ("SELECT COUNT(*) total FROM SOCIO WHERE estado = 'ACTIVO'", []),
                    "instrumentos": ("SELECT COUNT(*) total FROM INSTRUMENTO WHERE estado <> 'BAJA'", []),
                    "salas_disponibles": ("SELECT COUNT(*) total FROM SALA WHERE estado = 'ACTIVA'", []),
                    "reservas_confirmadas": ("SELECT COUNT(*) total FROM RESERVA WHERE estado = 'CONFIRMADA' AND eliminado_en IS NULL", []),
                    "prestamos_activos": ("SELECT COUNT(*) total FROM PRESTAMO WHERE estado = 'ACTIVO' AND eliminado_en IS NULL", []),
                }
            else:
                queries = {
                    "reservas_confirmadas": ("SELECT COUNT(*) total FROM RESERVA WHERE estado = 'CONFIRMADA' AND eliminado_en IS NULL AND socio_id = %s", [user_id]),
                    "prestamos_activos": ("SELECT COUNT(*) total FROM PRESTAMO WHERE estado = 'ACTIVO' AND eliminado_en IS NULL AND socio_id = %s", [user_id]),
                }

            for key, (sql, params) in queries.items():
                cursor.execute(sql, params)
                stats[key] = cursor.fetchone()["total"]
            
            sql_proximas = """
                SELECT r.id, r.fecha_inicio, s.nombre AS nombre_completo, sa.nombre AS sala_nombre, r.estado
                FROM RESERVA r
                JOIN SOCIO s ON s.id = r.socio_id
                JOIN SALA sa ON sa.id = r.sala_id
                WHERE r.fecha_inicio >= NOW()
                  AND r.estado = 'CONFIRMADA'
                  AND r.eliminado_en IS NULL
            """
            params_proximas = []
            if not is_admin:
                sql_proximas += " AND r.socio_id = %s"
                params_proximas.append(user_id)
            
            sql_proximas += " ORDER BY r.fecha_inicio LIMIT 5"
            
            cursor.execute(sql_proximas, params_proximas)
            stats["proximas_reservas"] = serialize_rows(cursor.fetchall())
            return ok(stats)
    finally:
        conn.close()


@app.route("/api/auditoria/reservas", methods=["GET"])
@require_admin
def audit_reservations():
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM AUDITORIA_RESERVA ORDER BY fecha_cambio DESC LIMIT 100")
            return ok(serialize_rows(cursor.fetchall()))
    finally:
        conn.close()


@app.route("/api/auditoria/prestamos", methods=["GET"])
@require_admin
def audit_loans():
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM AUDITORIA_PRESTAMO ORDER BY fecha_cambio DESC LIMIT 100")
            return ok(serialize_rows(cursor.fetchall()))
    finally:
        conn.close()


@app.route("/api/auditoria/instrumentos", methods=["GET"])
@require_admin
def audit_instruments():
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM AUDITORIA_INSTRUMENTO ORDER BY fecha_cambio DESC LIMIT 100")
            return ok(serialize_rows(cursor.fetchall()))
    finally:
        conn.close()


@app.route("/api/trigger_notifications", methods=["POST"])
@require_admin
def trigger_notifications():
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            now = datetime.now()
            target = now + timedelta(minutes=30)
            cursor.execute(
                """
                SELECT r.id, r.fecha_inicio, s.telefono, s.nombre
                FROM RESERVA r
                JOIN SOCIO s ON r.socio_id = s.id
                WHERE r.estado = 'CONFIRMADA'
                  AND r.eliminado_en IS NULL
                  AND r.fecha_inicio BETWEEN %s AND %s
                """,
                (now, target),
            )
            sent = 0
            for res in cursor.fetchall():
                primer_nombre = res["nombre"].split(" ")[0]
                if notify_whatsapp(res["telefono"], f"Recordatorio {primer_nombre}: tienes un ensayo a las {res['fecha_inicio']}."):
                    sent += 1
            return ok({"message": f"Notificaciones enviadas: {sent}"})
    finally:
        conn.close()

def check_overdue_loans():
    print("Ejecutando revisión de préstamos vencidos...")
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT p.id, p.socio_id, s.telefono, s.nombre, i.nombre as instrumento_nombre
                FROM PRESTAMO p
                JOIN SOCIO s ON p.socio_id = s.id
                JOIN INSTRUMENTO i ON p.instrumento_id = i.id
                WHERE p.estado = 'ACTIVO' AND p.fecha_limite < NOW() AND p.eliminado_en IS NULL
                """
            )
            vencidos = cursor.fetchall()
            
            for p in vencidos:
                cursor.execute("UPDATE PRESTAMO SET estado = 'VENCIDO', modificado_por = 'CRON' WHERE id = %s", (p["id"],))
                primer_nombre = p["nombre"].split(" ")[0]
                notify_whatsapp(
                    p["telefono"],
                    f"\u26a0\ufe0f ALERTA: Hola {primer_nombre}, el pr\u00e9stamo de tu instrumento '{p['instrumento_nombre']}' ha VENCIDO. Por favor descu\u00e9lvelo inmediatamente para evitar sanciones."
                )
                cursor.execute("""
                    INSERT IGNORE INTO MULTA (socio_id, prestamo_id, monto, motivo)
                    VALUES (%s, %s, 10.00, 'Préstamo vencido sin devolución')
                """, (p['socio_id'], p['id']))

            if vencidos:
                conn.commit()
                logger.info("Se actualizaron %d préstamos a estado VENCIDO.", len(vencidos))
    except Exception as exc:
        conn.rollback()
        logger.error("Error en Cron Job check_overdue_loans:", exc_info=True)
    finally:
        conn.close()

_notified_loan_ids: set = set()
_MAX_NOTIFIED_CACHE = 5000  # evitar crecimiento ilimitado en producción

try:
    from apscheduler.schedulers.background import BackgroundScheduler
    scheduler = BackgroundScheduler()
    scheduler.add_job(func=check_overdue_loans, trigger="interval", minutes=5)

    def check_upcoming_reservations():
        conn = get_db_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT r.id, r.fecha_inicio, s.nombre, s.telefono,
                           sa.nombre AS sala_nombre
                    FROM RESERVA r
                    JOIN SOCIO s ON s.id = r.socio_id
                    JOIN SALA sa ON sa.id = r.sala_id
                    WHERE r.estado = 'CONFIRMADA'
                      AND r.eliminado_en IS NULL
                      AND r.fecha_inicio BETWEEN NOW() AND DATE_ADD(NOW(), INTERVAL 65 MINUTE)
                      AND (r.notificado_1h IS NULL OR r.notificado_1h = 0)
                """)
                proximas = cursor.fetchall()
                for res in proximas:
                    primer_nombre = res['nombre'].split(' ')[0]
                    notify_whatsapp(
                        res['telefono'],
                        f"\u23f0 Recordatorio: Hola {primer_nombre}, tu reserva de *{res['sala_nombre']}* empieza en 1 hora ({res['fecha_inicio'].strftime('%H:%M') if isinstance(res['fecha_inicio'], datetime) else res['fecha_inicio']}). \u00a1No olvides llegar puntual!"
                    )
                    cursor.execute('UPDATE RESERVA SET notificado_1h = 1 WHERE id = %s', (res['id'],))
                if proximas:
                    conn.commit()
        except Exception as exc:
            logger.error("Error en reminder cron:", exc_info=True)
        finally:
            conn.close()

    def check_upcoming_loans():
        conn = get_db_connection()
        try:
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT p.id, p.fecha_limite, s.nombre, s.telefono,
                           i.nombre AS instrumento_nombre
                    FROM PRESTAMO p
                    JOIN SOCIO s ON p.socio_id = s.id
                    JOIN INSTRUMENTO i ON p.instrumento_id = i.id
                    WHERE p.estado = 'ACTIVO'
                      AND p.eliminado_en IS NULL
                      AND p.fecha_limite BETWEEN NOW() AND DATE_ADD(NOW(), INTERVAL 25 HOUR)
                """)
                proximos = cursor.fetchall()
                for p in proximos:
                    if p['id'] not in _notified_loan_ids:
                        primer_nombre = p['nombre'].split(' ')[0]
                        fecha_str = p['fecha_limite'].strftime('%d/%m/%Y') if hasattr(p['fecha_limite'], 'strftime') else str(p['fecha_limite'])
                        notify_whatsapp(
                            p['telefono'],
                            f"⏳ Hola {primer_nombre}, tu préstamo de *{p['instrumento_nombre']}* vence el {fecha_str}. Por favor devúelvelo a tiempo para evitar una multa."
                        )
                        _notified_loan_ids.add(p['id'])
                        if len(_notified_loan_ids) > _MAX_NOTIFIED_CACHE:
                            _notified_loan_ids.clear()
        except Exception as exc:
            logger.error("Error en reminder préstamos:", exc_info=True)
        finally:
            conn.close()

    scheduler.add_job(func=check_upcoming_reservations, trigger='interval', minutes=5)
    scheduler.add_job(func=check_upcoming_loans, trigger='interval', minutes=5)
    scheduler.add_job(func=_cleanup_token_blacklist, trigger='interval', hours=1)
    scheduler.start()
except ImportError:
    pass

@app.route('/api/public/info', methods=['GET'])
def public_info():
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, nombre, tipo, capacidad, equipamiento FROM SALA WHERE estado = 'ACTIVA' ORDER BY nombre")
            salas = serialize_rows(cursor.fetchall())
            cursor.execute("""
                SELECT id, nombre, fecha, lugar, descripcion, estado
                FROM EVENTO WHERE fecha >= NOW() AND estado IN ('PLANIFICADO','EN_PROGRESO')
                ORDER BY fecha LIMIT 5
            """)
            eventos = serialize_rows(cursor.fetchall())
            cursor.execute("SELECT COUNT(*) total FROM SOCIO WHERE estado = 'ACTIVO'")
            socios = cursor.fetchone()['total']
        return ok({'salas': salas, 'proximos_eventos': eventos, 'socios_activos': socios, 'club': 'Club de Música PUCESA'})
    finally:
        conn.close()


@app.route('/api/public/disponibilidad', methods=['GET'])
def public_disponibilidad():
    fecha_str = request.args.get('fecha')
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, nombre, tipo, capacidad FROM SALA WHERE estado = 'ACTIVA'")
            salas = cursor.fetchall()
            resultado = []
            for sala in salas:
                if fecha_str:
                    cursor.execute("""
                        SELECT COUNT(*) total FROM RESERVA
                        WHERE sala_id = %s AND DATE(fecha_inicio) = %s
                        AND estado IN ('CONFIRMADA','REPROGRAMADA') AND eliminado_en IS NULL
                    """, (sala['id'], fecha_str))
                    reservas_hoy = cursor.fetchone()['total']
                    sala['disponible_hoy'] = reservas_hoy == 0
                resultado.append(serialize_row(sala))
        return ok(resultado)
    finally:
        conn.close()


def _build_xlsx(title, subtitle, headers, col_widths, rows_data):
    """Genera un workbook de Excel con estilo profesional."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    # Paleta
    COLOR_HEADER_BG = '1E3A5F'   # azul marino
    COLOR_HEADER_FG = 'FFFFFF'
    COLOR_TITLE_BG  = '2563EB'   # azul brillante
    COLOR_ROW_ALT   = 'EFF6FF'   # azul muy claro
    COLOR_ROW_EVEN  = 'FFFFFF'
    COLOR_BORDER    = 'CBD5E1'

    thin = Side(style='thin', color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Título (fila 1, fusionada) ──────────────────────────────────────────
    n_cols = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font      = Font(name='Calibri', bold=True, size=16, color=COLOR_HEADER_FG)
    title_cell.fill      = PatternFill('solid', fgColor=COLOR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    # ── Subtítulo (fila 2) ──────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font      = Font(name='Calibri', italic=True, size=10, color='64748B')
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # ── Cabeceras (fila 3) ──────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = Font(name='Calibri', bold=True, size=11, color=COLOR_HEADER_FG)
        cell.fill      = PatternFill('solid', fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border
    ws.row_dimensions[3].height = 22

    # ── Datos (fila 4 en adelante) ──────────────────────────────────────────
    for row_idx, row_values in enumerate(rows_data, start=4):
        bg = COLOR_ROW_ALT if row_idx % 2 == 0 else COLOR_ROW_EVEN
        fill = PatternFill('solid', fgColor=bg)
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name='Calibri', size=10)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical='center', wrap_text=False)
        ws.row_dimensions[row_idx].height = 17

    # ── Anchos de columna ───────────────────────────────────────────────────
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Congelar cabeceras ──────────────────────────────────────────────────
    ws.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route('/api/reportes/reservas', methods=['GET'])
@require_admin
def export_reservas():
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            is_admin = request.user.get('rol') == 'ADMIN'
            user_id  = request.user.get('id')
            start    = request.args.get('start')
            end      = request.args.get('end')

            sql = """
                SELECT r.id, s.nombre AS socio, sa.nombre AS sala,
                       r.fecha_inicio, r.fecha_fin, r.estado,
                       COALESCE(r.observaciones, '') AS observaciones,
                       r.fecha_creacion
                FROM RESERVA r
                JOIN SOCIO s  ON s.id  = r.socio_id
                JOIN SALA  sa ON sa.id = r.sala_id
                WHERE r.eliminado_en IS NULL
            """
            params = []
            if not is_admin:
                sql += ' AND r.socio_id = %s'
                params.append(user_id)
            if start:
                sql += ' AND DATE(r.fecha_inicio) >= %s'
                params.append(start)
            if end:
                sql += ' AND DATE(r.fecha_fin) <= %s'
                params.append(end)
            sql += ' ORDER BY r.fecha_inicio DESC LIMIT 500'
            cursor.execute(sql, params)
            rows = cursor.fetchall()

        subtitle = f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")} · {len(rows)} registros'
        headers   = ['ID', 'Socio', 'Sala', 'Inicio', 'Fin', 'Duración (h)', 'Estado', 'Observaciones', 'Fecha registro']
        col_widths = [6, 24, 26, 22, 22, 14, 14, 32, 22]

        rows_data = []
        for r in rows:
            fi = r['fecha_inicio']
            ff = r['fecha_fin']
            duracion = round((ff - fi).total_seconds() / 3600, 2) if fi and ff else ''
            rows_data.append([
                r['id'], r['socio'], r['sala'],
                fi.strftime('%d/%m/%Y %H:%M') if fi else '',
                ff.strftime('%d/%m/%Y %H:%M') if ff else '',
                duracion,
                r['estado'], r['observaciones'],
                r['fecha_creacion'].strftime('%d/%m/%Y %H:%M') if r['fecha_creacion'] else '',
            ])

        buf = _build_xlsx('Reporte de Reservas', subtitle, headers, col_widths, rows_data)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=reservas.xlsx'}
        )
    finally:
        conn.close()


@app.route('/api/reportes/prestamos', methods=['GET'])
@require_admin
def export_prestamos():
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            is_admin = request.user.get('rol') == 'ADMIN'
            user_id  = request.user.get('id')
            start    = request.args.get('start')
            end      = request.args.get('end')

            sql = """
                SELECT p.id, s.nombre AS socio, i.nombre AS instrumento,
                       p.fecha_salida, p.fecha_limite, p.fecha_devolucion,
                       p.estado, COALESCE(p.observaciones, '') AS observaciones,
                       p.motivo
                FROM PRESTAMO p
                JOIN SOCIO      s ON s.id = p.socio_id
                JOIN INSTRUMENTO i ON i.id = p.instrumento_id
                WHERE p.eliminado_en IS NULL
            """
            params = []
            if not is_admin:
                sql += ' AND p.socio_id = %s'
                params.append(user_id)
            if start:
                sql += ' AND DATE(p.fecha_salida) >= %s'
                params.append(start)
            if end:
                sql += ' AND DATE(p.fecha_salida) <= %s'
                params.append(end)
            sql += ' ORDER BY p.fecha_salida DESC LIMIT 500'
            cursor.execute(sql, params)
            rows = cursor.fetchall()

        subtitle  = f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")} · {len(rows)} registros'
        headers   = ['ID', 'Socio', 'Instrumento', 'Motivo', 'Salida', 'Límite devolución', 'Devuelto el', 'Estado', 'Observaciones']
        col_widths = [6, 24, 26, 30, 22, 22, 22, 14, 32]

        rows_data = []
        for r in rows:
            fs  = r['fecha_salida']
            fl  = r['fecha_limite']
            fd  = r['fecha_devolucion']
            rows_data.append([
                r['id'], r['socio'], r['instrumento'],
                r.get('motivo', ''),
                fs.strftime('%d/%m/%Y %H:%M') if fs else '',
                fl.strftime('%d/%m/%Y %H:%M') if fl else '',
                fd.strftime('%d/%m/%Y %H:%M') if fd else '—',
                r['estado'], r['observaciones'],
            ])

        buf = _build_xlsx('Reporte de Préstamos', subtitle, headers, col_widths, rows_data)
        return Response(
            buf.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=prestamos.xlsx'}
        )
    finally:
        conn.close()


@app.route('/api/multas', methods=['GET'])
@require_auth
def get_multas():
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            is_admin = request.user.get('rol') == 'ADMIN'
            user_id = request.user.get('id')
            if is_admin:
                cursor.execute("""
                    SELECT m.*, s.nombre AS socio_nombre
                    FROM MULTA m JOIN SOCIO s ON s.id = m.socio_id
                    ORDER BY m.fecha_creacion DESC
                """)
            else:
                cursor.execute("""
                    SELECT m.*, s.nombre AS socio_nombre
                    FROM MULTA m JOIN SOCIO s ON s.id = m.socio_id
                    WHERE m.socio_id = %s ORDER BY m.fecha_creacion DESC
                """, (user_id,))
            return ok(serialize_rows(cursor.fetchall()))
    finally:
        conn.close()


@app.route('/api/multas', methods=['POST'])
@require_admin
def create_multa():
    data = parse_json()
    required = require_fields(data, ["socio_id", "motivo"])
    if required:
        return error(required)
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                INSERT INTO MULTA (socio_id, prestamo_id, reserva_id, monto, motivo)
                VALUES (%s, %s, %s, %s, %s)
            """, (data['socio_id'], data.get('prestamo_id'), data.get('reserva_id'),
                   data.get('monto', 5.00), data['motivo']))
            conn.commit()
            return ok({'message': 'Multa registrada', 'id': cursor.lastrowid}, 201)
    except Exception as exc:
        conn.rollback()
        logger.error("Internal error", exc_info=True)
        return error("Error interno del servidor", 500)
    finally:
        conn.close()


@app.route('/api/multas/<int:multa_id>/pagar', methods=['POST'])
@require_admin
def pay_multa(multa_id):
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                UPDATE MULTA SET estado = 'PAGADA', fecha_pago = NOW() WHERE id = %s
            """, (multa_id,))
            conn.commit()
            return ok({'message': 'Multa marcada como pagada'})
    except Exception as exc:
        conn.rollback()
        logger.error("Internal error", exc_info=True)
        return error("Error interno del servidor", 500)
    finally:
        conn.close()


@app.route("/api/statistics/loans", methods=["GET"])
@require_auth
def statistics_loans():
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            # Stats grouped by Tipo de Instrumento
            cursor.execute(
                """
                SELECT ti.nombre AS name, COUNT(p.id) AS value
                FROM PRESTAMO p
                JOIN INSTRUMENTO i ON i.id = p.instrumento_id
                JOIN TIPO_INSTRUMENTO ti ON ti.id = i.tipo_instrumento_id
                WHERE p.eliminado_en IS NULL
                GROUP BY ti.id, ti.nombre
                ORDER BY value DESC
                """
            )
            by_type = cursor.fetchall()

            # Stats grouped by Instrumento Específico
            cursor.execute(
                """
                SELECT i.nombre AS name, COUNT(p.id) AS value
                FROM PRESTAMO p
                JOIN INSTRUMENTO i ON i.id = p.instrumento_id
                WHERE p.eliminado_en IS NULL
                GROUP BY i.id, i.nombre
                ORDER BY value DESC
                """
            )
            by_instrument = cursor.fetchall()
            
            return ok({
                "by_type": by_type,
                "by_instrument": by_instrument
            })
    except Exception as exc:
        logger.error("Internal error", exc_info=True)
        return error("Error interno del servidor", 500)
    finally:
        conn.close()
if __name__ == "__main__":
    # Solo para depuración local. Producción usa gunicorn (ver Dockerfile).
    # 0.0.0.0 es necesario dentro de Docker para que nginx pueda alcanzar el contenedor.
    app.run(debug=False, host="0.0.0.0", port=5000)  # nosec B104
